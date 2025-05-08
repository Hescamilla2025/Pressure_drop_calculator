import streamlit as st
import numpy as np
import pandas as pd
import requests
from io import BytesIO
from openpyxl import load_workbook
from functions import friction_factor, pressure_loss, reynolds

st.title('Calculadora de caída de presión')

# Inicializar la tabla en session_state si no existe
if 'tabla' not in st.session_state:
    columnas = {
        'Flujo\n[m³/s]': [],
        'Velocidad\n[m/s]': [],
        'Densidad\n[kg/m³]': [],
        'Viscosidad\n[Pa·s]': [],
        'Factor de fricción\n[-]': [],
        'Caída de presión\n[Pa]': []
    }
    st.session_state.tabla = pd.DataFrame(columnas)

# Leer archivo Excel desde GitHub con requests y openpyxl
@st.cache_data
def cargar_excel_desde_github(url):
    response = requests.get(url)
    file = BytesIO(response.content)
    wb = load_workbook(file, data_only=True)
    return wb

# URLs de los archivos en GitHub
url_diametros = 'https://raw.githubusercontent.com/Hescamilla2025/Pressure_drop_calculator/main/primero.xlsx'
url_rugosidades = 'https://raw.githubusercontent.com/Hescamilla2025/Pressure_drop_calculator/main/segundo.xlsx'

# Cargar datos desde Excel
wb_diametros = cargar_excel_desde_github(url_diametros)
wb_rugosidades = cargar_excel_desde_github(url_rugosidades)

# Selección de material
st.subheader("Parámetros de tubería desde archivo Excel")
material = st.selectbox("Material de la tubería:", wb_diametros.sheetnames)  # Cambié keys() por sheetnames
ws_diametros = wb_diametros[material]

# Reconstruir DataFrame con la estructura de filas y columnas
df_material = pd.DataFrame(ws_diametros.values)
df_material.columns = df_material.iloc[0]
df_material = df_material.drop(0)
df_material.set_index(df_material.columns[0], inplace=True)  # Columna A como índice

cedula = st.selectbox("Cédula:", df_material.columns.tolist())
diametro_nominal = st.selectbox("Diámetro nominal:", df_material.index.tolist())

# Obtener diámetro
diametro = df_material.loc[diametro_nominal, cedula]  # en mm
st.write(f"📏 Diámetro interno: {diametro:.2f} mm")

# Rugosidad
usar_rugosidad_recomendada = st.checkbox("Usar rugosidad recomendada")
if usar_rugosidad_recomendada:
    ws_rugosidad = wb_rugosidades.active
    df_rug = pd.DataFrame(ws_rugosidad.values)
    df_rug.columns = ['Material', 'Rugosidad']
    df_rug.set_index('Material', inplace=True)
    rugosidad = df_rug.loc[material, 'Rugosidad']  # en mm
    st.write(f"🪵 Rugosidad recomendada: {rugosidad:.3f} mm")
else:
    rugosidad = float(st.text_input("Rugosidad de la tubería (mm):", value='0.046'))

# Formulario para agregar registro
st.subheader("Agregar registro")
with st.form("formulario"):
    flujo = float(st.text_input('Flujo en m³/s:', value='0.121'))
    densidad = float(st.text_input('Densidad del fluido (kg/m³):', value='1000'))
    viscosidad = float(st.text_input('Viscosidad del fluido (Pa·s):', value='0.00047'))
    longitud = float(st.text_input('Longitud de la tubería (m):', value='100'))

    # Cálculos
    area = np.pi * ((diametro / 2000) ** 2)  # Convertir mm a m, luego radio
    velocidad = flujo / area
    Re = reynolds(densidad, velocidad, diametro, viscosidad)
    friccion = friction_factor(diametro, Re, rugosidad)
    pressure_drop_m = pressure_loss(friccion, diametro, longitud, velocidad)
    pressure_drop_Pa = pressure_drop_m * densidad * 9.81

    # Mostrar resultados
    st.write(f'🔹 Velocidad: {velocidad:.3f} m/s')
    st.write(f'🔹 Número de Reynolds: {Re:.0f}')
    st.write(f'🔹 Factor de fricción: {friccion:.5f}')
    st.write(f'🔹 Caída de presión: {pressure_drop_m:.2f} m  |  {pressure_drop_Pa:.2f} Pa')
    agregar = st.form_submit_button("Agregar")

if agregar:
    nuevo = {
        'Flujo\n[m³/s]': flujo,
        'Velocidad\n[m/s]': velocidad,
        'Densidad\n[kg/m³]': densidad,
        'Viscosidad\n[Pa·s]': viscosidad,
        'Factor de fricción\n[-]': friccion,
        'Caída de presión\n[Pa]': pressure_drop_Pa
    }
    st.session_state.tabla.loc[len(st.session_state.tabla)] = nuevo
    st.success("✅ Registro agregado.")

# Mostrar tabla
st.subheader("Registros")
st.dataframe(st.session_state.tabla)

# Eliminar registros
st.subheader("Eliminar registros")
indices_a_eliminar = []
for i, row in st.session_state.tabla.iterrows():
    if st.checkbox(f"Eliminar fila {i}", key=f"del_{i}"):
        indices_a_eliminar.append(i)

if st.button("Eliminar seleccionados"):
    st.session_state.tabla.drop(indices_a_eliminar, inplace=True)
    st.session_state.tabla.reset_index(drop=True, inplace=True)
    st.success("🧹 Registros eliminados.")

# Mostrar suma total de la caída de presión
st.markdown(f"** Suma total de la caída de presión: {st.session_state.tabla['Caída de presión\n[Pa]'].sum():,.2f} Pa**")
